
# ============================================================
# HOW TO RUN:
# streamlit run "C:\Users\folde\financial_dashboard\financial_dashboard44.py"
# ============================================================

import streamlit as st
import yfinance as yf
import pandas as pd
import matplotlib.pyplot as plt
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ----------------------------
# Streamlit App Config
# ----------------------------
st.set_page_config(page_title="TayoTime Dashboard", layout="wide")
st.title("📊 TayoTime Invention and Investments - Apr 2026 - Revfin44")

# ----------------------------
# User Input
# ----------------------------
ticker_symbol = st.text_input(
    "Enter a stock ticker (e.g. AAPL, MSFT, TSLA):", 
    "AAPL"
).upper()
stock = yf.Ticker(ticker_symbol)

# ----------------------------
# Helper Functions
# ----------------------------
def safe_series(df, col):
    return df[col] if col in df.columns else pd.Series(0, index=df.index)

@st.cache_data
def load_financial_data(ticker):
    t = yf.Ticker(ticker)
    income = t.financials.T.copy()
    balance = t.balance_sheet.T.copy()
    cashflow = t.cashflow.T.copy()
    info = t.info or {}
    return income, balance, cashflow, info

def save_chart(fig):
    buffer = io.BytesIO()
    fig.savefig(buffer, format="png")
    buffer.seek(0)
    return buffer

# ----------------------------
# Current Price & Market Cap
# ----------------------------
try:
    hist_today = stock.history(period="1d")
    current_price = hist_today["Close"].iloc[-1]
    info = stock.info
    current_market_cap = info.get("marketCap", None)

    st.subheader(f"📈 Current {ticker_symbol} Price and Market Cap")
    st.metric(label="Latest Close", value=f"${current_price:,.2f}")
    if current_market_cap:
        st.metric(label="Market Cap", value=f"${current_market_cap/1e9:,.2f} B")
except Exception as e:
    st.warning(f"⚠️ Unable to fetch current stock price or market cap: {e}")





# ----------------------------
# Load Financials
# ----------------------------
try:
    income_df, balance_df, cashflow_df, info = load_financial_data(ticker_symbol)

    # Convert index to year
    for df in (income_df, balance_df, cashflow_df):
        df.index = pd.to_datetime(df.index).year

    available_years = sorted(set(income_df.index) & set(balance_df.index) & set(cashflow_df.index))
    st.markdown("### 🗓️ Available Financial Years")
    st.write(available_years)

    col1, col2 = st.columns(2)
    with col1:
        start_year = st.number_input(
            "Start Year", min_value=min(available_years), max_value=max(available_years), value=min(available_years)
        )
    with col2:
        end_year = st.number_input(
            "End Year", min_value=min(available_years), max_value=max(available_years), value=max(available_years)
        )

    def filter_years(df):
        return df[(df.index >= start_year) & (df.index <= end_year)]

    income_filtered = filter_years(income_df).round(2)
    balance_filtered = filter_years(balance_df).round(2)
    cashflow_filtered = filter_years(cashflow_df).round(2)

    # ----------------------------
    # Derived Metrics
    # ----------------------------
    metrics = pd.DataFrame(index=income_filtered.index)
    metrics["Revenue"] = safe_series(income_filtered, "Total Revenue")
    metrics["Net Income"] = safe_series(income_filtered, "Net Income")
    metrics["Free Cash Flow"] = safe_series(cashflow_filtered, "Free Cash Flow")
    metrics["Capex"] = safe_series(cashflow_filtered, "Capital Expenditure")
    metrics["Depreciation"] = safe_series(income_filtered, "Reconciled Depreciation")
    metrics["Gross Profit"] = safe_series(income_filtered, "Gross Profit")
    metrics["Gross Margin (%)"] = (metrics["Gross Profit"] / metrics["Revenue"].replace(0, pd.NA) * 100)
    metrics["Equity Book Value"] = safe_series(balance_filtered, "Stockholders Equity")
    metrics["Cash"] = safe_series(balance_filtered, "Cash And Cash Equivalents")
    metrics["Net Debt"] = safe_series(balance_filtered, "Net Debt")
    metrics["EPV @ 7%"] = (metrics["Free Cash Flow"] / 0.07) - metrics["Net Debt"]

# Graphing data 
    #metrics["EPV Absolute"] = metrics["EPV @ 7%"]- metrics["Equity Book Value"]
    #metrics["FV Absolute"] = metrics["EPV @ 7%"]







    # ----------------------------
    # ROIC Calculation
    # ----------------------------
    wacc = st.slider("WACC (%)", 5.0, 15.0, 9.0, 0.25)/100
    total_assets = safe_series(balance_filtered, "Total Assets")
    cash_eq = safe_series(balance_filtered, "Cash And Cash Equivalents")
    current_liab = safe_series(balance_filtered, "Total Current Liabilities")
    invested_capital = total_assets - cash_eq - current_liab
    nopat = safe_series(income_filtered, "Operating Income") * (1 - 0.21)

    if invested_capital.abs().sum() != 0:
        metrics["ROIC (%)"] = nopat / invested_capital.replace(0, pd.NA) * 100

    # ----------------------------
    # Historical Market Cap at Year-End
    # ----------------------------
    hist_full = stock.history(start=f"{start_year}-01-01", end=f"{end_year}-12-31")
    hist_full['Year'] = hist_full.index.year
    eoy_prices = hist_full.groupby('Year')['Close'].last()
    shares_outstanding = info.get("sharesOutstanding", None)

    if shares_outstanding:
        market_cap_hist = (eoy_prices * shares_outstanding).round(2)
        metrics["Market Cap"] = market_cap_hist
        metrics["Franchise Value"] = market_cap_hist
        metrics["EPV Required Growth %"] = ((metrics["Franchise Value"] / metrics["EPV @ 7%"]) - 1) * 100
        metrics["EPV Required Growth %"] = metrics["EPV Required Growth %"].replace(
            [float('inf'), -float('inf')], pd.NA
        ).round(2)

    metrics = metrics.round(2)
   

    # ----------------------------
    # SHOW CHARTS FIRST
    # ----------------------------
    charts = {}

# Book Value, EPV, Market Cap, Franchise Value Chart
    st.subheader("📊 Book Value, EPV, Market Cap, Franchise Value")
    combined_df = metrics[["Equity Book Value", "EPV @ 7%"]].copy()
    if "Market Cap" in metrics.columns:
        combined_df["Market Cap"] = metrics["Market Cap"]
        combined_df["Franchise Value"] = metrics["Franchise Value"]
    combined_df.fillna(0, inplace=True)

    fig, ax = plt.subplots(figsize=(12, 6))
    colors = {
        "Equity Book Value": "skyblue",
        "EPV @ 7%": "lightgreen",
        "Market Cap": "orange",
        "Franchise Value": "red"
    }
    combined_df.plot(kind="bar", ax=ax, color=[colors.get(c, "grey") for c in combined_df.columns])
    ax.set_title("Equity Book Value, EPV @ 7%, Market Cap, Franchise Value")
    ax.set_ylabel("USD")
    ax.set_xlabel("Year")
    ax.legend(title="Metric")

    if "EPV Required Growth %" in metrics.columns:
        for i, year in enumerate(metrics.index):
            growth = metrics.loc[year, "EPV Required Growth %"]
            if pd.notna(growth):
                max_val = combined_df.iloc[i].max()
                ax.text(i, max_val * 1.02, f"{growth:.1f}%", ha='center', fontsize=10, fontweight='bold', color='black')

    fig.tight_layout()
    st.pyplot(fig)
    charts["Book_EPV_MarketCap_Franchise"] = save_chart(fig)
    
# ROIC Chart
    if "ROIC (%)" in metrics.columns:
        st.subheader("🏭 Return on Invested Capital (ROIC)")
        fig, ax = plt.subplots()
        metrics["ROIC (%)"].dropna().plot(kind="bar", ax=ax, color="purple")
        ax.axhline(wacc*100, linestyle="--", color="red", label="WACC")
        ax.set_ylabel("ROIC (%)")
        ax.set_title("ROIC vs WACC")
        ax.legend()
        st.pyplot(fig)
        charts["ROIC"] = save_chart(fig)
    else:
        st.warning("ROIC is not meaningful for this business model.")

    # Gross Margin Chart
    fig, ax = plt.subplots()
    metrics["Gross Margin (%)"].dropna().sort_index().plot(kind="bar", ax=ax, color="skyblue")
    ax.set_title("Gross Margin (%)")
    ax.set_ylabel("Percent")
    ax.set_xlabel("Year")
    fig.tight_layout()
    st.pyplot(fig)
    charts["Gross Margin (%)"] = save_chart(fig)

    # Revenue, FCF, Net Income Chart
    st.subheader("📊 Revenue, Free Cash Flow, Net Income")
    combined_income_df = metrics[["Revenue", "Free Cash Flow", "Net Income"]].fillna(0)
    fig, ax = plt.subplots(figsize=(10, 6))
    combined_income_df.plot(kind="bar", ax=ax)
    ax.set_title("Revenue, Free Cash Flow, Net Income")
    ax.set_ylabel("USD")
    ax.set_xlabel("Year")
    ax.legend(title="Metric")
    fig.tight_layout()
    st.pyplot(fig)
    charts["Revenue_FCF_NetIncome"] = save_chart(fig)

    

    # ----------------------------
    # THEN SHOW TABLES
    # ----------------------------
    st.subheader("📈 Derived Financial Metrics")
    #st.dataframe(metrics) = metrics.loc[:, list(metrics.columns)]
    st.dataframe(metrics)


    st.subheader("📄 Income Statement")
    st.dataframe(income_filtered)

    st.subheader("📄 Balance Sheet")
    st.dataframe(balance_filtered)

    st.subheader("📄 Cash Flow Statement")
    st.dataframe(cashflow_filtered)

    # ----------------------------
    # Excel Export
    # ----------------------------
    try:
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            income_filtered.to_excel(writer, sheet_name="Income Statement")
            balance_filtered.to_excel(writer, sheet_name="Balance Sheet")
            cashflow_filtered.to_excel(writer, sheet_name="Cash Flow")
            metrics.to_excel(writer, sheet_name="Metrics")

        wb = load_workbook(excel_buffer)
        for sheet_name, chart_img in charts.items():
            ws = wb.create_sheet(title=sheet_name + " Chart")
            img = XLImage(chart_img)
            ws.add_image(img, "B2")

        final_buffer = io.BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)

        st.download_button(
            label="📥 Download Excel (Data + Charts)",
            data=final_buffer,
            file_name=f"{ticker_symbol}_financials_{start_year}-{end_year}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        st.error(f"Could not create Excel export: {e}")

except Exception as e:
    st.error(f"⚠️ Error fetching or processing data: {e}")